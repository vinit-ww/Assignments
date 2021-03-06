import xlsxwriter
import openpyxl
class Excel(object):

    def create(self):
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook('Expenses01.xlsx')
        worksheet = workbook.add_worksheet()

        # Some data we want to write to the worksheet.
        expenses = (
            ['Rent', 1000],
            ['Gas',   100],
            ['Food',  300],
            ['Gym',    50],
            ['Ront', 1000],
            ['Gs',   100],
            ['Fod',  3090],
            ['Gy',    590],
            ['ent', 10000],
            ['Ga',   1080],
            ['Food',  3800],
            ['Gm',    560],
            ['Rent', 10050],
            ['Gas',   100],
            ['ood',  3040],
            ['Gym',    50],
            ['Rent', 100670],
            ['Gas',   103240],
            ['Food',  30014],
            ['Gym',    150],
            ['Rent', 1031400],
            ['Gas',   10430],
            ['Food',  33400],
            ['Gym',    53450],
            
            )

        # Start from the first cell. Rows and columns are zero indexed.
        row = 0
        col = 0

        # Iterate over the data and write it out row by row.
        for item, cost in (expenses):
            worksheet.write(row,col,item)
            worksheet.write(row, col + 1, cost)
            row += 1
        workbook.close()    

    def modify(self):
        wb=openpyxl.load_workbook('Expenses01.xlsx')
        c=wb.get_sheet_names()
        sheet=wb.get_sheet_by_name(c[0])
        d=sheet['A1']
        print d.value
        for i in range(1,8,2):
            print(sheet.cell(row=i,column=1).value)
                        


obj=Excel()
#obj.create()
obj.modify()